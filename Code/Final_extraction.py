import re
import pandas as pd
from openpyxl import load_workbook

RAW_FILE = r"C:\Drive_d\Python\F-AI\T7\Invoice_Output.xlsx"
TEMPLATE_FILE = r"C:\Drive_d\Python\F-AI\T4\Output Template.xlsx"
OUTPUT_FILE = r"C:\Drive_d\Python\F-AI\T7\Invoice_Filled.xlsx"

def num(x):
    try:
        return float(re.sub(r"[^\d.]", "", str(x)))
    except:
        return 0.0

def norm(c):
    return re.sub(r"[^a-z0-9]", "", str(c).lower())

def find_col(cols, keys):
    for c in cols:
        if all(k in c for k in keys):
            return c
    return None

def grab(p, t):
    m = re.search(p, t, re.I | re.S)
    return m.group(1).strip() if m else ""

# ================= TABLE =================
df = pd.read_excel(RAW_FILE, sheet_name="Table", dtype=str)
df = df.dropna(axis=1, how="all")

df.columns = df.iloc[0]
df = df[1:].reset_index(drop=True)
df.columns = [norm(c) for c in df.columns]

df = df[
    ~df.iloc[:, 0].astype(str).str.contains("total|amount", case=False)
]

c_desc = find_col(df.columns, ["item", "description"])
c_mrp = find_col(df.columns, ["mrp"])
c_disc = find_col(df.columns, ["discount"])
c_qty = find_col(df.columns, ["qty"])
c_taxable = find_col(df.columns, ["taxable"])
c_cgst = find_col(df.columns, ["cgst", "inr"])
c_sgst = find_col(df.columns, ["sgst", "inr"])
c_total = find_col(df.columns, ["total"])

total_amount = df[c_total].apply(num).sum() if c_total else 0
total_tax = (
    df[c_cgst].apply(num).sum() if c_cgst else 0
) + (
    df[c_sgst].apply(num).sum() if c_sgst else 0
)

# ================= HEADERS =================
text = []
for i in range(1, 10):
    try:
        h = pd.read_excel(RAW_FILE, sheet_name=f"Header_{i}", header=None)
        text.append(" ".join(h.astype(str).values.flatten()))
    except:
        pass

full_text = " ".join(text)

meta = {
    "billing_address": grab(r"Sold By\s*(.*?)GSTIN", full_text),
    "shipping_address": grab(r"Sold By\s*(.*?)GSTIN", full_text),
    "invoice_type": "Tax Invoice",
    "order_number": grab(r"Order Id\s*(\d+)", full_text),
    "invoice_number": grab(r"Order Id\s*(\d+)", full_text),
    "order_date": grab(r"(\d{2}-[A-Za-z]{3}-\d{4})", full_text),
    "invoice_date": grab(r"(\d{2}-[A-Za-z]{3}-\d{4})", full_text),
    "seller_name": "Zomato Hyperpure Private Limited",
    "seller_gst": grab(r"GSTIN\s*(\d{2}[A-Z0-9]+)", full_text),
    "seller_address": grab(r"Sold By\s*(.*?)GSTIN", full_text),
    "place_of_supply": grab(r"Place of Supply\s*(\w+)", full_text),
    "total_tax": total_tax,
    "total_amount": total_amount,
}

# ================= FILL TEMPLATE =================
wb = load_workbook(TEMPLATE_FILE)
ws_meta, ws_items = wb.worksheets[:2]

for r in ws_meta.iter_rows(min_row=2, max_col=2):
    if r[0].value in meta:
        r[1].value = meta[r[0].value]

ws_items.delete_rows(2, ws_items.max_row)

for i, r in df.iterrows():
    ws_items.append([
        i+1,
        r[c_desc] if c_desc else "",
        num(r[c_mrp]) if c_mrp else 0,
        num(r[c_disc]) if c_disc else 0,
        num(r[c_qty]) if c_qty else 0,
        num(r[c_taxable]) if c_taxable else 0,
        "",
        "CGST+SGST",
        (num(r[c_cgst]) + num(r[c_sgst])) if c_cgst and c_sgst else 0,
        num(r[c_total]) if c_total else 0,
    ])

wb.save(OUTPUT_FILE)
print("✅ FINAL TEMPLATE GENERATED:", OUTPUT_FILE)
