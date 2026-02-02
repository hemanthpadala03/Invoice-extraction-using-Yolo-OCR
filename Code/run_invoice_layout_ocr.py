import os
import cv2
import easyocr
import pandas as pd
from ultralytics import YOLO
from openpyxl import Workbook

# =====================================================
# PATH CONFIG
# =====================================================
BASE_DIR = r"C:\Drive_d\Python\F-AI\T7"

MODEL_PATH = r"C:\Drive_d\Python\F-AI\T7\runs\detect\invoice_layout\yolov8s_layout_v14\weights\best.pt"
IMAGE_PATH = r"C:\Drive_d\Python\F-AI\T7\Input\images\val\ForwardInvoice_ORD43073428557_page_1.png"
OUTPUT_EXCEL = r"C:\Drive_d\Python\F-AI\T7\Invoice_Output.xlsx"

# =====================================================
# LOAD MODELS
# =====================================================
model = YOLO(MODEL_PATH)
reader = easyocr.Reader(['en'], gpu=False)

# =====================================================
# LOAD IMAGE
# =====================================================
img = cv2.imread(IMAGE_PATH)
if img is None:
    raise RuntimeError("Image not found")

H, W, _ = img.shape

# =====================================================
# YOLO INFERENCE
# =====================================================
results = model(img, conf=0.35, iou=0.5)[0]

headers = []
tables = []

for box in results.boxes:
    cls_id = int(box.cls[0])
    cls_name = model.names[cls_id]
    x1, y1, x2, y2 = map(int, box.xyxy[0])

    if cls_name == "header_block":
        headers.append((x1, y1, x2, y2))
    elif cls_name == "table":
        tables.append((x1, y1, x2, y2))

# =====================================================
# SORT HEADERS (TOP → BOTTOM)
# =====================================================
headers = sorted(headers, key=lambda b: b[1])

# =====================================================
# OCR HEADERS ONE BY ONE
# =====================================================
header_texts = {}

for idx, (x1, y1, x2, y2) in enumerate(headers, start=1):
    crop = img[y1:y2, x1:x2]
    ocr = reader.readtext(crop, detail=0, paragraph=True)
    header_texts[f"Header_{idx}"] = "\n".join(ocr).strip()

# =====================================================
# TABLE EXTRACTION (UNCHANGED LOGIC)
# =====================================================
table_data = []

if tables:
    # take largest table
    table_box = max(tables, key=lambda b: (b[2]-b[0]) * (b[3]-b[1]))
    x1, y1, x2, y2 = table_box
    table_img = img[y1:y2, x1:x2]

    th, tw, _ = table_img.shape

    gray = cv2.cvtColor(table_img, cv2.COLOR_BGR2GRAY)
    _, bw = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU
    )

    # ---------- Horizontal lines ----------
    h_kernel = cv2.getStructuringElement(
        cv2.MORPH_RECT, (int(tw * 0.4), 1)
    )
    h_lines = cv2.morphologyEx(
        bw, cv2.MORPH_OPEN, h_kernel, iterations=2
    )

    h_contours, _ = cv2.findContours(
        h_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
    )

    h_positions = [0, th - 1]
    for cnt in h_contours:
        x, y, w_line, h_line = cv2.boundingRect(cnt)
        if w_line > tw * 0.6:
            h_positions.append(y)

    h_positions = sorted(set(h_positions))

    # ---------- Vertical lines ----------
    v_kernel = cv2.getStructuringElement(
        cv2.MORPH_RECT, (1, int(th * 0.4))
    )
    v_lines = cv2.morphologyEx(
        bw, cv2.MORPH_OPEN, v_kernel, iterations=2
    )

    v_contours, _ = cv2.findContours(
        v_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
    )

    v_positions = [0, tw - 1]
    for cnt in v_contours:
        x, y, w_line, h_line = cv2.boundingRect(cnt)
        if h_line > th * 0.5:
            v_positions.append(x)

    v_positions = sorted(set(v_positions))

    # ---------- CELL OCR ----------
    for i in range(len(h_positions) - 1):
        row = []
        for j in range(len(v_positions) - 1):
            y1, y2 = h_positions[i], h_positions[i + 1]
            x1, x2 = v_positions[j], v_positions[j + 1]

            if (y2 - y1) < 15 or (x2 - x1) < 15:
                row.append("")
                continue

            cell = table_img[y1:y2, x1:x2]
            ocr = reader.readtext(cell, detail=0, paragraph=True)
            row.append(" ".join(ocr).strip())

        table_data.append(row)

# =====================================================
# SAVE TO EXCEL
# =====================================================
wb = Workbook()

# Save headers (Header-1 untouched, Header-2 normalized)
for idx, ((name, text), box) in enumerate(zip(header_texts.items(), headers)):
    if idx == 0:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(title=name)

    x1, y1, x2, y2 = box
    box_w = x2 - x1
    box_h = y2 - y1

    # Heuristic:
    # wide + short → single-line header (Header-2 case)
    if box_w / max(box_h, 1) > 6:
        clean_text = " ".join(text.split())
        ws.append([clean_text])
    else:
        # Header-1 behavior (UNCHANGED)
        ws.append([text])


# Save table
ws_table = wb.create_sheet("Table")
df = pd.DataFrame(table_data)

for r in df.itertuples(index=False):
    ws_table.append(list(r))

wb.save(OUTPUT_EXCEL)

print("✅ DONE")
print("Saved:", OUTPUT_EXCEL)
